"""
Вспомогательные функции
"""
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import aiofiles
from config import EMAIL_TO, SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, RESPONSES_DIR


async def load_bot_structure(structure_file: Path) -> Dict:
    """Загружает структуру бота из JSON файла"""
    async with aiofiles.open(structure_file, 'r', encoding='utf-8') as f:
        content = await f.read()
        return json.loads(content)


def determine_group(user_answers: Dict) -> Optional[int]:
    """
    Определяет группу на основе ответов пользователя
    
    Логика основана на первом вопросе:
    - "С отношениями к еде и телу" -> группа 15 (Строимость)
    - "С деньгами и ощущением стабильности" -> группа 16 (Финансы)
    - "С уверенностью в себе" -> группа 17 (Самооценка)
    - "С отношениями с близкими" -> группа 18 (Отношения)
    - "С привычками от которых сложно отказаться" -> группа 19 (Негативные привычки)
    """
    first_answer = user_answers.get('question_1', '').lower()
    
    # Более точное определение по ключевым словам
    if any(word in first_answer for word in ['еде', 'тело', 'переедани', 'диет', 'строимость']):
        return 15
    elif any(word in first_answer for word in ['деньга', 'финанс', 'стабильност']):
        return 16
    elif any(word in first_answer for word in ['уверенност', 'самооценк', 'сомнени']):
        return 17
    elif any(word in first_answer for word in ['отношени', 'близк', 'довери']):
        return 18
    elif any(word in first_answer for word in ['привычк', 'зависимост', 'алкогол']):
        return 19
    
    return None


async def save_user_responses(user_id: int, username: str, answers: Dict, group_id: Optional[int]):
    """Сохраняет ответы пользователя в Excel файл"""
    # Создаем директорию если её нет
    RESPONSES_DIR.mkdir(exist_ok=True)
    
    # Создаем или открываем файл Excel
    excel_file = RESPONSES_DIR / f'responses_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    if excel_file.exists():
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ответы пользователей"
        
        # Заголовки
        headers = ['Дата', 'ID пользователя', 'Username', 'Вопрос 1', 'Вопрос 2', 'Вопрос 3', 'Определенная группа']
        ws.append(headers)
        
        # Форматирование заголовков
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Добавляем строку с данными
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        user_id,
        username or 'N/A',
        answers.get('question_1', ''),
        answers.get('question_2', ''),
        answers.get('question_3', ''),
        group_id or 'Не определена'
    ]
    ws.append(row)
    
    # Сохраняем файл
    wb.save(excel_file)
    
    return excel_file


async def send_email_with_excel(excel_file: Path):
    """Отправляет Excel файл на email"""
    if not all([EMAIL_TO, SMTP_USER, SMTP_PASSWORD]):
        return False
    
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = EMAIL_TO
        msg['Subject'] = f"Ответы пользователей - {datetime.now().strftime('%Y-%m-%d')}"
        
        body = f"Прикреплен файл с ответами пользователей за {datetime.now().strftime('%Y-%m-%d')}"
        msg.attach(MIMEText(body, 'plain'))
        
        # Прикрепляем файл
        with open(excel_file, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {excel_file.name}'
        )
        msg.attach(part)
        
        # Отправляем
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        text = msg.as_string()
        server.sendmail(SMTP_USER, EMAIL_TO, text)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Ошибка отправки email: {e}")
        return False


async def send_to_moderator_telegram(bot, user_id: int, username: str, answers: Dict, group_id: Optional[int]):
    """Отправляет ответы пользователя модератору в Telegram"""
    from config import MODERATOR_ID
    
    if not MODERATOR_ID:
        return False
    
    message = f"""📋 Новые ответы пользователя:

👤 Пользователь: @{username or 'N/A'} (ID: {user_id})
📅 Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

❓ Вопрос 1: {answers.get('question_1', 'N/A')}
❓ Вопрос 2: {answers.get('question_2', 'N/A')}
❓ Вопрос 3: {answers.get('question_3', 'N/A')}

🎯 Определенная группа: {group_id or 'Не определена'}
"""
    
    try:
        await bot.send_message(MODERATOR_ID, message)
        return True
    except Exception as e:
        print(f"Ошибка отправки сообщения модератору: {e}")
        return False

